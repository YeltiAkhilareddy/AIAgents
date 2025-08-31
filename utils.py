from requests.auth import HTTPBasicAuth
from django.utils.timezone import now
from datetime import datetime
from django.db.models import Prefetch
from .models import TicketAssignment
from collections import defaultdict
import pytz
import time
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from .models import Ticket, Team  # Import your models
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.response import Response
from django.core.exceptions import ObjectDoesNotExist
from django.utils.dateparse import parse_datetime
from datetime import datetime, timezone, timedelta
import json
from presalesApp.models import Project
from users.models import User
from typing import List
from pydantic import BaseModel, ValidationError
from django.db.models import Q
from .models import Category, ChatConvo, Module, Prompt, Team
import requests
import os
from django.conf import settings
from pydantic import BaseModel
import requests
import io
import re
import openpyxl
from ticketing.models import Ticket, TicketAssignment, TrackingRecord

groq_api_key = settings.GROQ_API_KEY
embedding_api_key = settings.EMBEDDING_API_KEY
embedding_api_url = settings.EMBEDDING_API_URL
openai_api_key = settings.OPENAI_API_KEY
openai_api_url = settings.OPENAI_API_URL

client = None
if groq_api_key:
    from groq import Groq
    client = Groq(api_key=groq_api_key)

# -----------------------------------------------Clean----Code------------------------------------------------------------


def fetch_memory_contexts(chatwindow_id: int, user, limit: int = 10) -> str:
    try:
        chat_history = (
            ChatConvo.objects.filter(
                Q(chatwindow_id=chatwindow_id) & Q(user=user)
            )
            .order_by("-timestamp")
            [:limit]
        )

        if not chat_history.exists():
            return "No previous context found."

        memory_context = "\n".join(
            [f"user_message: {chat.user_query} Response: {chat.response}" for chat in chat_history]
        )

        return memory_context

    except Exception as e:
        print(f"Error fetching memory context: {e}")
        return f"Error: {str(e)}"


def generate_unique_id(user_id):
    timestamp = datetime.now().strftime("%H%M%S")
    unique_id = int(f"{timestamp}")
    return unique_id


def get_ai_responses(api_url=openai_api_url, api_key=openai_api_key, user_message=None, memory_context=None, chatwindow_id=None, user=None, temperature=0.5):
    prompt_instance = Prompt.objects.filter(
        name="CSA").order_by('-created_at').first()
    if not prompt_instance:
        raise RuntimeError("Prompt content not found in database.")

    headers = {
        "api-key": api_key,
        "Content-Type": "application/json"
    }

    try:
        memory_context = fetch_memory_contexts(
            chatwindow_id=int(chatwindow_id), user=user)
    except Exception:
        memory_context = ""
    ids = generate_unique_id(user.id)
    memory_context_text = f"\n\nYou have Memory and you Can Recall the Conversation if user Ask, Previous Conversation: {memory_context}" if memory_context else ""

    prompt = {
        "role": "system",
        "content": f"{memory_context_text} \n {prompt_instance.content} \n\n ticket id should be this Number {ids} \n\n ticket_creation timesatmp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                   "Ensure your response is **strictly** valid JSON, without any additional text, headers, or formatting. Do not include explanations or introductory statements."
    }

    messages = [prompt, {"role": "user", "content": user_message}]

    payload = {
        "messages": messages,
        "temperature": temperature
    }

    response = requests.post(api_url, headers=headers, json=payload)

    if response.status_code == 200:
        data = response.json()
        ai_response = data.get("choices", [{}])[0].get(
            "message", {}).get("content", "{}")

        try:
            ai_response = ai_response.strip().strip("```json").strip("```").strip()
            parsed_response = json.loads(ai_response)

            if not isinstance(parsed_response, dict):
                parsed_response = {
                    "error": "AI response is not a valid JSON object"}

        except json.JSONDecodeError:
            parsed_response = {"error": "Failed to parse AI response"}
        return parsed_response, parsed_response.get("response", "")

    else:
        return {"success": False, "error": f"Error: {response.status_code}, {response.text}"}


def create_sap_ticket_from_context_Experimentals(memory_context: str) -> dict:
    category_list: List[str] = list(
        Category.objects.values_list("name", flat=True))
    application_modules_list: List[str] = list(
        Module.objects.values_list("name", flat=True))
    team_list: List[str] = list(
        Team.objects.values_list("name", flat=True))

    ticket_type_list: List[str] = [
        "Incident", "Service"]

    sap_ticket_details = generate_sap_ticket_details(
        memory_context=memory_context,
        category_list=category_list,
        application_modules_list=application_modules_list,
        ticket_type_list=ticket_type_list,
        teams_list=team_list
    )

    return sap_ticket_details.dict()


def postprocess_tickets(user_id, response_data, project, tic_id):
    try:
        india_tz = pytz.timezone("Asia/Kolkata")
        ticket_details = response_data.get("ticket_details", [])

        for ticket in ticket_details:
            module_id = ticket["module"]
            category_id = ticket["category"]

            module = Module.objects.filter(id=module_id).first()
            category = Category.objects.filter(category_id=category_id).first()

            if not module:
                return {"success": False, "error": f"Module with ID {module_id} does not exist."}
            if not category:
                return {"success": False, "error": f"Category with ID {category_id} does not exist."}
            try:
                user = User.objects.get(id=user_id)
            except ObjectDoesNotExist:
                return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

            try:
                project_instance = Project.objects.get(id=project)
            except ObjectDoesNotExist:
                return Response({"error": "Project not found."}, status=status.HTTP_404_NOT_FOUND)

            ticket_creation = Ticket.objects.create(
                ticket_id=ticket["ticket_id"],
                user=user,
                title=ticket["title"],
                description=ticket["description"],
                module=module,
                category=category,
                project=project_instance,
                priority=ticket["priority"],
                project_type=ticket["ticket_type"],
                status=ticket.get("status", "Assigned"),
            )
            api_url = settings.OPENAI_API_URL
            api_key = settings.OPENAI_API_KEY
            ticket_preview = get_suitable_team(tic_id, api_url, api_key)
            mail = "natarjn@gmail.com"
            send_email(
                to_email=mail,
                subject="New Ticket Assigned to you",
                template_name="ticketAssignedToTeam",
                context={
                    "team_name": mail.split('@')[0],
                    "team": "Service Desk Agent",
                    "ticket_id": tic_id,
                    "ticket_subject": ticket["title"],
                    "ticket_description": ticket["description"],
                    "ticket_module": module,
                    "ticket_status": "Assigned",
                    "assigned_by": "Customer Service Agent",
                    "ticket_priority": ticket["priority"],
                    "assigned_at": datetime.now(india_tz).strftime("%Y-%m-%d-%H:%M"),
                    "company_name": "Niveda AI",
                }
            )
            send_email(
                to_email=user.email,
                subject="Your Support Ticket Has Been Created",
                template_name="ticketCreationMail",
                context={
                    "user_name": user.name,
                    "ticket_id": tic_id,
                    "ticket_subject": ticket["title"],
                    "ticket_priority": ticket["priority"],
                    "ticket_status": "Assigned",
                    "team_name": ticket_preview['team_assigned'],
                    "company_name": "Niveda AI",
                }
            )

            team_instance = Team.objects.get(
                name=ticket_preview['team_assigned'])
            ticket_Assign = TicketAssignment.objects.create(
                ticket=ticket_creation,
                assigned_to=team_instance,
                status='Assigned',
                comments=f"Service Desk Agent assigned this ticket to {ticket_preview['team_assigned']} Team"
            )
            comment = generate_ticket_routing_responseExpo(
                ticket_creation.ticket_id)
            TicketAssignment.objects.filter(
                ticket_assignment_id=ticket_Assign.ticket_assignment_id).update(comments=comment)
            TicketAssignment.objects.create(
                ticket=ticket_creation,
                assigned_to=team_instance,
                status='In Progress',
                comments=comment
            )
            send_email(
                to_email=ticket_preview['email'],
                subject="New Ticket Assigned to Your Team",
                template_name="ticketAssignedToTeam",
                context={
                    "team_name": ticket_preview['email'].split('.')[0] + ' Team',
                    "team": ticket_preview['team_assigned'],
                    "ticket_id": tic_id,
                    "ticket_subject": ticket["title"],
                    "ticket_description": ticket["description"],
                    "ticket_module": module,
                    "ticket_status": "Assigned",
                    "assigned_by": "Service Desk Agent",
                    "ticket_priority": ticket["priority"],
                    "assigned_at": datetime.now(india_tz).strftime("%Y-%m-%d-%H:%M"),
                    "company_name": "Niveda AI",
                }
            )

            ticket_instance = Ticket.objects.get(
                ticket_id=ticket_creation.ticket_id)
            error_message = ticket_instance.description

            solutions = get_solutionagent(
                ticket_id=ticket_creation.ticket_id)
            Ticket.objects.filter(
                ticket_id=ticket_creation.ticket_id).update(solution=solutions['solution'])

        return {"success": True, "message": "Tickets saved successfully"}

    except Exception as e:
        print(f"Error: {str(e)}")
        return {"success": False, "error": str(e)}


def get_solutionagent(ticket_id, api_url=openai_api_url, api_key=openai_api_key):
    try:
        ticket = Ticket.objects.get(ticket_id=ticket_id)
        error_message = ticket.description
    except Ticket.DoesNotExist:
        return {"success": False, "error": f"Ticket with ID {ticket_id} not found."}

    headers = {
        "api-key": api_key,
        "Content-Type": "application/json"
    }

    prompt_content = f"""
    You are an **SAP Expert Consultant** 🏆 specializing in troubleshooting SAP errors.
    Your task is to provide a **detailed, step-by-step resolution** for the given SAP error.

    📌 **Response Formatting Rules:**
    ✅ Always respond **only** in **Markdown format** (no JSON).
    ✅ Use **headings, bold text, bullet points, and numbered lists** for clarity.
    ✅ Keep responses **structured, concise, and actionable**.

    ---
    ## 🔍 **1. Identify the Issue**
    - ❓ **Error Meaning:** What does this error generally mean in SAP?
    - 🏢 **Affected Module:** Identify the relevant SAP module (e.g., **MM, SD, FI, ABAP, Basis, PP**).
    - ⚠️ **Common Scenarios:** When does this error typically occur? Provide examples.

    ---
    ## 🛠 **2. Step-by-Step Solution**
    📝 Provide a structured **step-by-step guide** to resolve the issue:
    1️⃣ **Step 1:** First action to take (e.g., check system settings, T-Code, configuration).
    2️⃣ **Step 2:** Next steps to investigate or fix the issue.
    3️⃣ **Step 3:** Continue with actionable steps until the issue is resolved.
    🔄 **Alternative Fixes:** If multiple solutions exist, list them in priority order.

    ---
    ## ✅ **3. Example Fix (If Applicable)**
    📌 Provide a **real-world SAP scenario** illustrating how this issue was resolved.
    💡 Include **SAP transaction codes (T-Codes)** where relevant.

    ---
    ## 🛡️ **4. Preventive Measures**
    - 🔄 **Best Practices:** Steps to prevent this issue in the future.
    - 🎓 **Training & Awareness:** Educate users on avoiding this error.
    - 🔔 **System Checks & Automation:** Use monitoring tools to detect and prevent similar issues.

    ---
    ### **🛑 Error**: {error_message}

    """

    payload = {
        "messages": [
            {"role": "system", "content": prompt_content},
            {"role": "user", "content": error_message}
        ],
        "temperature": 0.7,
        "max_tokens": 2048
    }

    try:
        response = requests.post(api_url, headers=headers, json=payload)
        response.raise_for_status()

        try:
            data = response.json()
        except json.JSONDecodeError:
            return {"success": False, "error": "Invalid JSON response from AI API."}
        ai_response = data.get("choices", [{}])[0].get(
            "message", {}).get("content", "").strip()
        if not ai_response:
            return {"success": False, "error": "AI response is empty or malformed."}

        return {"success": True, "solution": ai_response}

    except requests.RequestException as e:
        return {"success": False, "error": f"Request failed: {str(e)}"}


def generate_ticket_routing_responseExpo(ticket_id: int):
    """
    Fetches ticket data, analyzes it, and generates a conversational markdown response
    from the perspective of a Service Desk Agent.
    """

    try:
        ticket_data = fetch_ticket_and_team_detailsExpo(ticket_id)

        if not ticket_data:
            return "**Error:** Ticket not found."

        groq_api_key = os.getenv("GROQ_API_KEY") or getattr(
            settings, "GROQ_API_KEY_3", None)
        if not groq_api_key:
            raise ValueError("GROQ_API_KEY is missing or invalid.")

        client = Groq(api_key=groq_api_key)
        prompt_content = f"""
            You are Amelia, a **Service Desk Agent** responsible for analyzing incoming tickets, determining the best support team, and ensuring SLA compliance.

            Your response should be:
            ✅ **Conversational** – Talk directly to the customer.
            ✅ **Professional** – Keep it clear, structured, and reassuring.
            ✅ **Markdown-Formatted** – Use headings, lists, and emphasis to make it easy to read.

            ---

            ## 💬 Service Desk Agent Response
            Hi there, I’ve reviewed your ticket and assigned it to the best team to resolve this issue. Here’s what you need to know:

            ---

            ## 🎫 Ticket Summary
            - **Ticket ID:** {ticket_data['ticket_id']}
            - **Title:** {ticket_data['title']}
            - **Issue:** {ticket_data['description']}
            - **Category:** {ticket_data['category']}
            - **Module:** {ticket_data['module']}
            - **Project:** {ticket_data['project']}
            - **Ticket Type:** {ticket_data['project_type']}

            ---

            ## 🛠 Team Selection & Assignment
            Based on my analysis, the best team to handle this issue is the **{ticket_data['team_assigned']} team**. They specialize in resolving issues related to this module and will begin working on it shortly.

            ---

            ## 📢 What Happens Next?
            - **🕒 SLA Compliance:**
            - This ticket has been marked as **{ticket_data['priority']} priority**, meaning the target resolution time is **{ticket_data['sla']} hours**.
            - **🚀 Next Steps:**
            - The **{ticket_data['team_assigned']} team** will start working on the issue.
            - If any additional details are needed, they will reach out to you.
            - If the issue is unresolved within the SLA, it will be escalated.
            - **📡 Tracking & Updates:**
            - I will keep you informed of any progress.
            - Any status changes or escalations will be communicated to you promptly.

            ---

            ## 📌 Summary
            Your ticket **(ID: {ticket_data['ticket_id']})** has been successfully assigned to the **{ticket_data['team_assigned']} team**, and they are now working on it.

            **📍 SLA Target:** {ticket_data['sla']} hours

            🚀 I’ll keep you updated! 😊
            """
        chat_completion = client.chat.completions.create(
            messages=[{"role": "system", "content": prompt_content}],
            model="llama-3.3-70b-versatile",
            temperature=0,
            stream=False,
            response_format={"type": "text"},
        )

        response_text = chat_completion.choices[
            0].message.content if chat_completion.choices else "**Error:** LLM response invalid."

        return response_text

    except Exception as e:
        return f"**Error:** An issue occurred while generating the response - {e}"


def fetch_ticket_and_team_detailsExpo(ticket_id):
    try:
        ticket = Ticket.objects.get(ticket_id=ticket_id)
        priority_sla_mapping = {
            "High": 4,
            "Medium": 8,
            "Low": 12
        }
        sla_hours = priority_sla_mapping.get(ticket.priority, 4)
        assigned_team = TicketAssignment.objects.filter(ticket=ticket).first()
        ticket_data = {
            "ticket_id": ticket.ticket_id,
            "title": ticket.title,
            "description": ticket.description,
            "priority": ticket.priority,
            "sla": sla_hours,
            "status": ticket.status,
            "category": ticket.category.name,
            "module": ticket.module.name,
            "project": ticket.project.name,
            "project_type": ticket.project_type,
            "team_assigned": assigned_team.assigned_to.name if assigned_team else "Not Assigned"
        }

        return ticket_data

    except Ticket.DoesNotExist:
        return None


class SAPResponse(BaseModel):
    application_module: str
    category: str
    priority: str
    ticket_title: str
    ticket_description: str
    ticket_type: str
    team_assigned: str


def generate_sap_ticket_details(memory_context: str, category_list: List[str], application_modules_list: List[str], ticket_type_list: List[str], teams_list: List[str]) -> SAPResponse:
    """
    Generates SAP ticket details based on conversation memory and predefined lists.

    Parameters:
        memory_context (str): Previous conversation history.
        category_list (List[str]): List of allowed categories.
        application_modules_list (List[str]): List of SAP application modules.
        ticket_type_list (List[str]): List of valid ticket types.

    Returns:
        SAPResponse: Structured SAP ticket details.
    """
    try:
        groq_api_key = settings.GROQ_API_KEY_2
        client = Groq(api_key=groq_api_key)
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "system",
                    "content": (
                        f"You are Amelia, an expert SAP customer service agent.\n\n"
                        "Analyze the memory context provided and generate SAP ticket details based on the issue described.\n\n"
                        f"**Memory Context:**\n{memory_context}\n\n"
                        "**Ensure the following conditions:**\n"
                        "- Select an appropriate `application_module` from this predefined list:\n"
                        f"{application_modules_list}\n"
                        "- Choose a valid `category` from:\n"
                        f"{category_list}\n"
                        "- Assign a suitable `priority` from ['High', 'Medium', 'Low'] based on issue severity.\n"
                        "- Generate a clear and meaningful `ticket_title` summarizing the issue.\n"
                        "- Provide a detailed `ticket_description` explaining the problem Only.\n"
                        "- Choose an appropriate `ticket_type` from:\n"
                        f"{ticket_type_list}\n\n"
                        "Ensure your response is **strictly valid JSON** following this structure:\n"
                        "select  an appropriate `team_assigned` from this predefined teams list:\n"
                        f"{teams_list}\n"
                        f"{json.dumps(SAPResponse.model_json_schema(), indent=2)}"
                    ),
                }
            ],
            model="llama-3.3-70b-versatile",
            temperature=0,
            stream=False,
            response_format={"type": "json_object"},
        )

        if not chat_completion.choices or not chat_completion.choices[0].message.content:
            raise RuntimeError(
                "Received an empty or invalid response from the model.")

        return SAPResponse.model_validate_json(chat_completion.choices[0].message.content)

    except ValidationError as e:
        print(f"Validation error: {e}")
        raise RuntimeError(
            "The API response does not match the expected SAPResponse schema.")

    except Exception as e:
        print(f"Error generating SAP ticket details: {e}")
        raise RuntimeError(
            "An error occurred while generating the SAP ticket details.")


def process_ticket_response(ticket_id: int):
    """
    Fetches ticket details from the API, processes them, and formats a professional response in Markdown format.
    Also checks SLA compliance based on the assigned SLA time.
    """
    try:
        # Fetch ticket data from API
        api_url = f"http://13.127.202.241:8000/api/v1/aitickets/{ticket_id}/"
        response = requests.get(api_url)
        if response.status_code != 200:
            return "**Error:** Unable to fetch ticket data from API."

        ticket_data = response.json()
        if not ticket_data.get("success") or "data" not in ticket_data:
            return "**Error:** Invalid ticket data received."

        data = ticket_data["data"]
        ticket_id = data.get("ticket_id", "N/A")
        title = data.get("title", "N/A")
        description = data.get("description", "N/A")
        priority = data.get("priority", "N/A")
        project_type = data.get("project_type", "N/A")
        module_name = data.get("module_name", "N/A")
        category_name = data.get("category_name", "N/A")
        created_at = data.get("created_at", "N/A")
        assignments = data.get("ticket_assignments", [])

        if not assignments:
            return "**Error:** No team assignments found for this ticket."

        latest_assignment = max(
            assignments, key=lambda x: x.get('assigned_at', ''))
        assigned_team = latest_assignment.get("assigned_to", "N/A")
        current_status = latest_assignment.get("status", "N/A")
        assigned_at = latest_assignment.get("assigned_at", "N/A")
        sla_target_hours = {'High': 4, 'Medium': 8, 'Low': 12}.get(
            priority, 8)

        created_dt = datetime.fromisoformat(created_at.replace(
            "Z", "+00:00")) if created_at != "N/A" else None
        assigned_dt = datetime.fromisoformat(assigned_at.replace(
            "Z", "+00:00")) if assigned_at != "N/A" else None
        current_dt = datetime.now(timezone.utc)

        sla_breached = False
        sla_deadline = None
        if created_dt:
            sla_deadline = created_dt + timedelta(hours=sla_target_hours)
            sla_breached = current_dt > sla_deadline

        response_text = f"""
        ## 🕵️ Responsive Agent Report
        Hey! I've been monitoring your ticket, and here’s the latest update on its status:

        ## 📋 Ticket Summary
        - **Ticket ID:** {ticket_id}
        - **Title:** {title}
        - **Issue:** {description}
        - **Category:** {category_name}
        - **Module:** {module_name}
        - **Ticket Type:** {project_type}
        - **Priority:** {priority}
        - **Created At:** {created_at}
        
        ## 🔍 Monitoring & Current Status
        - **Assigned Team:** {assigned_team}
- **Current Status:** {current_status}        - **Assigned At:** {assigned_at} 📅
        
        ## ⏳ SLA Monitoring
        - **SLA Target:** {sla_target_hours} hours ⏳
        - **SLA Deadline:** {sla_deadline.strftime('%Y-%m-%d %H:%M:%S UTC') if sla_deadline else 'N/A'}
        - **SLA Status:** {'🚨 Breached' if sla_breached else '✅ Within SLA'}
        
        ## 🚀 Next Steps
        - The **{assigned_team}** team will start working on the issue.
        - If any additional details are needed, they will reach out to you.
        - If the issue is unresolved within the SLA, it will be escalated.
        - Any status changes or escalations will be communicated to you promptly.
        
        ## 📌 Summary
        Your ticket **(ID: {ticket_id})** has been successfully assigned to the **{assigned_team} team**, and they are now working on it.

        **📍 SLA Target:** {sla_target_hours} hours

        I'll be keeping a close watch on this and will notify you about any changes!
        """
        return response_text.strip()

    except Exception as e:
        return f"**Error:** An issue occurred while processing the ticket - {e}"


def get_all_teams():
    teams = Team.objects.all().values("team_id", "name", "email")
    if not teams:
        return {"success": False, "error": "No teams found in the database."}
    teams_list = [{"team_id": team["team_id"], "team_name": team["name"],
                   "email": team["email"]} for team in teams]
    return {"success": True, "teams": teams_list}


def get_ticket_data(ticket_id):
    ticket = get_object_or_404(Ticket, ticket_id=ticket_id)

    # Fetch related assignments
    assignments = ticket.assignments.all().values(
        "ticket_assignment_id",
        "assigned_to",
        "status",
        "comments",
        "assigned_at"
    )

    return {
        "ticket_id": ticket.ticket_id,
        "title": ticket.title,
        "description": ticket.description,
        "module": ticket.module,
        "project": ticket.project,
        "priority": ticket.priority,
        "category": ticket.category,
        "project_type": ticket.project_type,
        "status": ticket.status,
        "created_at": ticket.created_at,
        "updated_at": ticket.updated_at,
        "solution": ticket.solution,
        "assignments": list(assignments),
    }


def is_sla_breached(priority, created_at):
    sla = {"low": 12, "medium": 8, "high": 4}
    if priority not in sla:
        raise ValueError(
            "Invalid priority. Choose from 'low', 'medium', or 'high'.")

    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at)

    if created_at.tzinfo is None:
        created_at = created_at.replace(tzinfo=timezone.utc)

    now = datetime.now(timezone.utc)
    sla_deadline = created_at + timedelta(hours=sla[priority])

    return now > sla_deadline


def Response_Agent_function(ticket_id, api_url=openai_api_url, api_key=openai_api_key):
    try:
        ticket = Ticket.objects.get(ticket_id=ticket_id)
        error_message = ticket.description
    except Ticket.DoesNotExist:
        return {"success": False, "error": f"Ticket with ID {ticket_id} not found."}

    ticket_data = get_ticket_data(ticket_id)

    # Convert datetime objects to ISO format strings
    def format_datetime(dt):
        return dt.isoformat() if isinstance(dt, datetime) else dt

    ticket_data["created_at"] = format_datetime(
        ticket_data.get("created_at"))
    ticket_data["updated_at"] = format_datetime(
        ticket_data.get("updated_at"))

    for assignment in ticket_data.get("assignments", []):
        assignment["assigned_at"] = format_datetime(
            assignment.get("assigned_at"))

    headers = {
        "api-key": api_key,
        "Content-Type": "application/json"
    }

    assignments_data = '\n'.join(
        [f"| {a['assigned_to']} | {a['status']} | {repr(a['comments'])} | {a['assigned_at']} |"
         for a in ticket_data.get("assignments", [])]
    )

    breached = is_sla_breached(
        ticket_data.get("priority").lower(), ticket_data["created_at"])
    sla_status = f"SLA Breached: {breached}"

    prompt_content = f"""
        🤖 You are a Responsive Ticket Analysis Agent
        Your role is to analyze SAP incident tickets, track SLA compliance, and monitor assignment escalations. Provide a structured, precise, and insightful report based strictly on the given details.

        🔍 1. Ticket Details
        Ticket ID: {ticket_data.get("ticket_id", "N/A")}
        Title: {ticket_data.get("title", "N/A")}
        Description: {ticket_data.get("description", "N/A")}
        Module: {ticket_data.get("module", "N/A")}
        Category: {ticket_data.get("category", "N/A")}
        Priority: {ticket_data.get("priority", "N/A")}
        Project Type: {ticket_data.get("project_type", "N/A")}
        Status: {ticket_data.get("status", "N/A")}
        Created At: {ticket_data.get("created_at", "N/A")}
        Last Updated: {ticket_data.get("updated_at", "N/A")}
        ⏳ 2. SLA Tracking
        🚨 SLA Compliance Check:

        SLA Breached: {sla_status}
        Action:
        If SLA is breached, flag it as 🔴 SLA BREACHED and send a notification to the responsible team.
        If SLA is within limits, mark it as 🟢 SLA OK.
        📌 3. Assignment & Escalation Tracking
        Assignment History
        📜 Ticket Assignment Log:
        {assignments_data} # Don't show raw assignment data

        (Example format: strictly follow)

        [Timestamp] Assigned to: [assignee] (By: [assigner])
        [Timestamp] Reassigned to: [next_assignee]
        Escalation Flow Analysis
        🔄 Escalation Path:

        The ticket has been assigned from:
        Customer Service Agent → Service Desk Agent
        Service Desk Agent → Configuration Team
        📢 Escalation Insights:

        If the ticket has multiple escalations, determine if they were necessary or inefficient.
        If the ticket is stuck in "Assigned" status beyond an acceptable duration, flag it for escalation to higher management.

        📊 4. Conclusion & Recommendations
        📌 Final Analysis:

        Root cause of delay (if any): [root_cause]
        Efficiency of resolution process: [resolution_efficiency]
        Recommended actions to prevent similar delays: [recommendations]
        ✅ Keep your response concise, structured, and strictly within the provided data and include all four Sections.
        ❌ DO NOT add extra sections or information beyond the instructions.
    """

    payload = {
        "messages": [
            {"role": "system", "content": prompt_content},
            {"role": "user", "content": error_message}
        ],
        "temperature": 0.7,
        "max_tokens": 2048
    }

    try:
        response = requests.post(api_url, headers=headers, json=payload)
        response.raise_for_status()

        try:
            data = response.json()
        except json.JSONDecodeError:
            return {"success": False, "error": "Invalid JSON response from AI API."}

        ai_response = data.get("choices", [{}])[0].get(
            "message", {}).get("content", "").strip()
        if not ai_response:
            return {"success": False, "error": "AI response is empty or malformed."}

        return {"success": True, "solution": ai_response}

    except requests.RequestException as e:
        return {"success": False, "error": f"Request failed: {str(e)}"}


# ---------------------------------Compliance Agent--------------------------------------------

def validate_fetch_excel_from_s3(s3_url):
    """Fetch an Excel file from an S3 URL and return a file-like object."""
    try:
        response = requests.get(s3_url, stream=True)
        if response.status_code == 200:
            return io.BytesIO(response.content)
        else:
            raise Exception(
                f"Failed to retrieve file from S3. Status Code: {response.status_code}")
    except Exception as e:
        raise Exception(f"S3 Fetch Error: {str(e)}")


def validate_serial_number_exists_in_excel(file_stream, serial_number):
    """Check if a given serial number exists in the Excel file."""
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        sheet = wb.active

        # Adjust min_row as needed
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if row[1] and str(row[1]).strip() == str(serial_number):
                return True

        return False
    except Exception as e:
        raise Exception(f"Excel Processing Error: {str(e)}")


def validate_extract_text_from_excel(file_stream):
    """Extract raw text from an Excel file, handling unstructured data."""
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        extracted_data = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = []

            for row in sheet.iter_rows(values_only=True):
                row_data = [
                    str(cell) if cell is not None else "" for cell in row]
                sheet_data.append(" | ".join(row_data))

            extracted_data[sheet_name] = "\n".join(sheet_data)

        return extracted_data
    except Exception as e:
        raise Exception(f"Excel Text Extraction Error: {str(e)}")


def validate_track_serial_number(file_stream, serial_number, api_url, api_key, prompt_template, temperature=0.5):
    """Fetch Excel data, extract information, and send it to Azure OpenAI for processing."""
    try:
        raw_text = validate_extract_text_from_excel(file_stream)

        headers = {
            "api-key": api_key,
            "Content-Type": "application/json"
        }

        # Format the prompt with the provided serial number
        formatted_prompt = prompt_template.replace(
            "{serial_number}", serial_number)

        # OpenAI request payload
        payload = {
            "messages": [
                {"role": "system", "content": f"{formatted_prompt}"},
                {"role": "user", "content": json.dumps(raw_text, indent=2)}
            ],
            "temperature": temperature
        }

        # Send request to OpenAI API
        response = requests.post(api_url, headers=headers, json=payload)

        if response.status_code == 200:
            data = response.json()
            ai_response = data.get("choices", [{}])[0].get(
                "message", {}).get("content", "{}")

            # Debugging: Print raw AI response
            print("🔍 AI Raw Response:", ai_response)

            try:
                ai_response = re.sub(r"```json|```", "", ai_response).strip()

                print("ai-response", ai_response)

            except json.JSONDecodeError:
                structured_response = {"error": "Failed to parse AI response"}

            return {
                "success": True,
                "response": ai_response
            }
        else:
            return {
                "success": False,
                "error": f"Error: {response.status_code}, {response.text}"
            }
    except Exception as e:
        return {
            "success": False,
            "error": f"Processing Error: {str(e)}"
        }


def get_suitable_team(ticket_id, api_url, api_key):
    """
    Assigns a ticket to the most suitable team using GPT-4.0 from Azure.
    """
    try:
        # Fetch ticket details
        ticket = Ticket.objects.get(ticket_id=ticket_id)
        teams = Team.objects.all()

        # Format teams
        teams_info = "\n".join([
            f"Team Name: {team.name}, Team ID: {team.team_id}, Email: {team.email}"
            for team in teams
        ])

        # Format ticket details
        ticket_details = f"""
        Title: {ticket.title}
        Description: {ticket.description}
        Module: {ticket.module}
        Project: {ticket.project}
        Priority: {ticket.priority}
        Category: {ticket.category}
        Project Type: {ticket.project_type}
        Status: {ticket.status}
        """

        # Construct GPT prompt
        prompt = f"""
        Given the following ticket details, determine the most suitable team from the available options.

        Ticket Details:
        {ticket_details}

        Available Teams:
        {teams_info}

        Respond ONLY in valid JSON format with 'team_id', 'team_name', and 'email'.
        """

        # Azure API request
        headers = {
            "Content-Type": "application/json",
            "api-key": api_key
        }

        payload = {
            "messages": [
                {"role": "system", "content": "You are an AI that assigns tickets to the most appropriate team."},
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 100
        }

        response = requests.post(
            api_url, headers=headers, json=payload, timeout=10)
        response.raise_for_status()
        response_json = response.json()

        # Default value
        team_info = {"error": "No valid response from GPT model"}

        if "choices" in response_json and response_json["choices"]:
            raw_response = response_json["choices"][0]["message"]["content"]
            clean_json = re.sub(r"```json|```", "", raw_response).strip()

            try:
                team_info = json.loads(clean_json)
                team_info["application_module"] = ticket.module.name if ticket.module else None
                team_info["category"] = ticket.category.name if ticket.category else None
                team_info["priority"] = ticket.priority
                team_info["ticket_title"] = ticket.title
                team_info["ticket_description"] = ticket.description
                team_info["ticket_type"] = ticket.project_type
                team_info["team_assigned"] = team_info.get("team_name")
            except json.JSONDecodeError:
                team_info = {"error": "Invalid JSON response from the model"}

    except requests.Timeout:
        team_info = {"error": "Request timed out"}
    except requests.RequestException as e:
        team_info = {"error": f"API request failed: {str(e)}"}
    except ObjectDoesNotExist:
        team_info = {"error": f"Ticket with ID {ticket_id} not found"}
    except Exception as e:
        team_info = {"error": f"Unexpected error: {str(e)}"}

    return team_info


def send_email(to_email, subject, template_name, context):
    html_message = render_to_string(f"emails/{template_name}.html", context)
    email = EmailMessage(
        subject=subject,
        body=html_message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[to_email],
    )
    email.content_subtype = "html"

    try:
        email.send()
    except Exception as e:
        print(f"Failed to send email: {e}")


def send_sla_breach_notifications():
    breached_tickets = Ticket.objects.filter(escalation_notified=False)

    for ticket in breached_tickets:
        if is_sla_breached(ticket.priority.lower(), ticket.created_at):
            latest_assignment = TicketAssignment.objects.filter(
                ticket=ticket).order_by("-assigned_at").first()

            if not latest_assignment:
                continue

            assigned_team = latest_assignment.assigned_to
            team_email = assigned_team.email if assigned_team.email else "default_support@yourdomain.com"

            context = {
                "ticket_title": ticket.title,
                "ticket_id": ticket.ticket_id,
                "priority": ticket.priority,
                "created_at": ticket.created_at,
                "status": ticket.status,
                "assigned_team": assigned_team.name
            }

            subject = f"SLA Breach Alert: {ticket.title} (ID: {ticket.ticket_id})"
            send_email(team_email, subject, "sla_breach", context)
            ticket.escalation_notified = True
            ticket.save()


def process_ticket_assignments():
    ticket_ids = TicketAssignment.objects.filter(
        email_triggered=False
    ).values_list('ticket_id', flat=True).distinct()

    assignments_to_update = []

    for ticket_id in ticket_ids:
        assignments = TicketAssignment.objects.filter(
            ticket_id=ticket_id
        ).order_by('assigned_at')

        previous_assignment = None

        for assignment in assignments:
            if not previous_assignment or not previous_assignment.email_triggered:
                previous_assignment = assignment

            if assignment.email_triggered:
                continue

            status_changed = assigned_to_changed = False

            if previous_assignment and previous_assignment.ticket_assignment_id != assignment.ticket_assignment_id:
                status_changed = assignment.status != previous_assignment.status
                assigned_to_changed = assignment.assigned_to_id != previous_assignment.assigned_to_id

                if status_changed or assigned_to_changed:
                    context = {
                        "recipient_name": assignment.assigned_to.name if assigned_to_changed else assignment.ticket.user.name,
                        "ticket_id": ticket_id,
                        "status": assignment.status,
                        "previous_status": previous_assignment.status if previous_assignment else "N/A",
                        "assigned_to": assignment.assigned_to.name,
                        "previous_assigned_to": previous_assignment.assigned_to.name if previous_assignment else "N/A",
                        "is_status_changed": status_changed,
                        "is_assignment_changed": assigned_to_changed,
                        "current_year": now().year,
                        "company_name": "Niveda Business AI Solutions Pvt Ltd"
                    }

                    send_email(
                        to_email=assignment.ticket.user.email,
                        subject=f"Ticket #{ticket_id} Update Notification",
                        template_name="ticket_update",
                        context=context
                    )

                    if assigned_to_changed:
                        send_email(
                            to_email=assignment.assigned_to.email,
                            subject=f"New Ticket Assigned: #{ticket_id}",
                            template_name="ticket_update",
                            context=context
                        )

                    assignments_to_update.append(assignment)

            previous_assignment = assignment

    TicketAssignment.objects.filter(ticket_assignment_id__in=[
        a.ticket_assignment_id for a in assignments_to_update
    ]).update(email_triggered=True)

    return assignments_to_update


def fetch_cleaned_json(endpoint, username, password):
    """
    Fetches JSON data from an OData API endpoint and returns it.
    """
    try:
        response = requests.get(
            endpoint, auth=HTTPBasicAuth(username, password))
        response.raise_for_status()
        json_data = response.json()

        return json_data

    except requests.exceptions.RequestException as e:
        return {"error": str(e)}
    except ValueError as e:
        return {"error": f"JSON Parsing Error: {str(e)}"}
